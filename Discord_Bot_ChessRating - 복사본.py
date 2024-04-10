import discord
from discord.ext import commands
from openpyxl import load_workbook, Workbook

TOKEN = 
bot = commands.Bot(command_prefix='!',intents=discord.Intents.all())

c_name = 1
c_point = 2
c_win = 3
c_lose = 4

default_point = 1000

wb = load_workbook("userDB.xlsx")
ws = wb.active

def signup(_name):
    _row = ws.max_row + 1
    
    ws.cell(row=_row, column=c_name, value=_name)
    ws.cell(row=_row, column=c_point, value = default_point)
    ws.cell(row=_row, column=c_win, value = 0)
    ws.cell(row=_row, column=c_lose, value = 0)
    wb.save("userDB.xlsx")

def checkName(_name):
    for row in range(2, ws.max_row+1):
        if ws.cell(row,1).value == _name:
            return False
        else:
            continue

    return True

def delete():
    ws.delete_rows(2,ws.max_row)
    wb.save("userDB.xlsx")

def userInfo(_name):
    if not checkName(_name):
        for row in range(2, ws.max_row+1):
            if ws.cell(row, 1).value == _name:
                return ws.cell(row,1).value, ws.cell(row,2).value, ws.cell(row,3).value, ws.cell(row,4).value
    else:
    	return None, None, None, None

def usercheck(_winner, _loser):
    winner_row = 0
    loser_row = 0

    for row in range(2, ws.max_row+1):
        if ws.cell(row, 1).value == _winner:
            winner_row = row
        elif ws.cell(row, 1).value == _loser:
            loser_row = row

    if winner_row == 0 and loser_row == 0:
        return 1
    elif winner_row == 0:
        return 2
    elif loser_row == 0:
        return 3
    
def pointupdown(_winner, _loser):
    winner_row = 0
    loser_row = 0

    for row in range(2, ws.max_row+1):
        if ws.cell(row, 1).value == _winner:
            winner_row = row
        elif ws.cell(row, 1).value == _loser:
            loser_row = row

    ws.cell(winner_row, 2).value += (int)((ws.cell(loser_row, 2).value + ws.cell(winner_row, 3).value * 1000) * 0.1)
    ws.cell(winner_row, 4).value = 0
    if ws.cell(winner_row, 3).value >= 0:
        ws.cell(winner_row, 3).value += 1

    ws.cell(loser_row, 2).value -= (int)((ws.cell(loser_row, 2).value + ws.cell(loser_row, 4).value * 1000) * 0.1)
    ws.cell(loser_row, 3).value = 0
    if ws.cell(loser_row, 4).value >= 0:
        ws.cell(loser_row, 4).value -= 1

    wb.save("userDB.xlsx")
    
           

################################비동기#############################################
@bot.event
async def on_ready():
    print(f'Login bot: {bot.user}')

@bot.command()
async def 도움말(ctx):
    await ctx.send("회원가입 : 유저의 계정이 생성됩니다. ex) !회원가입")
    await ctx.send("랭킹 : 유저의 랭킹이 표시됩니다. ex) !랭킹")
    await ctx.send("승패 [승자] [패자]: 유저의 전적을 갱신합니다. ex) !승패 jeongjaeyeong. 주찬경")


@bot.command()
async def 회원가입(ctx):
    if checkName(ctx.author.name):
        signup(ctx.author.name)
        await ctx.send("회원가입이 완료되었습니다!")
    else:
        await ctx.send("이미 가입된 사용자입니다.")

@bot.command()
async def reset(ctx):
    delete()

@bot.command()
async def 랭킹(ctx):
    name, point, wp, lp = userInfo(ctx.author.name)

    if name == None or point == None:
        await ctx.send("등록되지 않은 사용자입니다.""!회원가입"" 후 사용해주세요")
    else:
        embeded = discord.Embed(title = "유저 정보", description=ctx.author.name, color = 0x62D0F6)
        embeded.add_field(name="Rating", value = point)
        embeded.add_field(name="연승", value = wp)
        embeded.add_field(name="연패", value = lp)

        await ctx.send(embed=embeded)

@bot.command()
async def 승패(ctx, arg1, arg2):
    if usercheck(arg1, arg2)==1:
        await ctx.send("둘 다 등록되지 않은 사용자를 입력하였습니다.")
    elif usercheck(arg1, arg2)==2:
        await ctx.send("첫 번째 사용자가 등록되지 않은 사용자를 입력하였습니다.")
    elif usercheck(arg1, arg2)==3:
        await ctx.send("두 번째 사용자가 등록되지 않은 사용자를 입력하였습니다.")
    else:
        pointupdown(arg1, arg2)
        await ctx.send("Rating이 수정되었습니다!")


bot.run(TOKEN)